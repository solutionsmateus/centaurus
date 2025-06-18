from openpyxl import Workbook
from openpyxl import load_workbook

wb = Workbook()
workbook = load_workbook(filename="C:\\Users\\Pedro.gomes\\Grupo Mateus\\Guilherme Soares de Macedo - Analise Encartes Concorrencia\\Analise de Produtos e Concorrencia.xlsx")
workbook.sheetnames


sheet = workbook.active
sheet

ws = wb.active
ws.title = "Encartes"


headers = [
    "Empresa", "Validade", "Categoria do Produto", "Data Inicio", "Data Fim",
    "Campanha", "Produto", "Preço", "App", "Estado"
]
ws.append(headers)


produtos = ["C:\\Users\\Pedro.gomes\\Grupo Mateus\\Guilherme Soares de Macedo - Analise Encartes Concorrencia\\Analise de Produtos e Concorrencia.xlsx"]


for rows in sheet.iter_rows(max_col=100000,min_col=1, min_row=1, max_row= 10.0000):
     ws.append("C:\\Users\\Pedro.gomes\\Grupo Mateus\\Desktop\\data\\image_reader.csv")

for produto in produtos:
    ws.append(produto)

file_path = "C:\\API\\EncMark\\docs"
wb.save(file_path)

file_path
